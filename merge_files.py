# This program merges files from Google Drive's folders, by user's choice.
# 1. The user provides a country which he wants to merge its data (0 for france, 1 for germany, 2 for uk and 3 for us).
# After this step, the relevant "all_data" file is being downloaded
# NOTE: we assume that there is only one "ALL" file per country.
# 2. The user provides a number represents which merge should be done:
# if a sentiment query was done, make sure that after the manual verification the file was moved from "for verification"
# folder to "after verification" folder, and enter 0.
# otherwise, make sure that the file is in the "deltas" folder and enter 1.
# 3. The relevant file will be downloaded and merged with the relevant "all_data" file.
# NOTE - If there are tweets that has been collected twice (in the DB), they will be removed from the first DF.
#        if a post appears in both DFs but has been collected only once (may happen when merging different queries), and
#        has different values in the same column, the conflict will raise as error, with the relevant details.
#        handle errors like this manually, by editing one of the files.
# 4. If the merging was finished without any errors, the original files will be moved to the archive folder, and the
# merged file will be uploaded to the Drive. All local copies are being removed automatically.
import os
from pathlib import Path
import IDs
import time
import pandas as pd
import drive_functions as d
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from tweet_classifier import TaskName, ClassificationTask
# this copy should handle posts that were collected twice by deleting one of them (the copy from the first df)
# other duplicates would raise as errors.


# read the input file
def read_file(file_path):
    return pd.read_excel(file_path)


def combine_values(row, col):
    if pd.isnull(row[col + '_df1']):
        return row[col + '_df2']
    elif pd.isnull(row[col + '_df2']):
        return row[col + '_df1']
    elif row[col + '_df1'] == row[col + '_df2']:
        return row[col + '_df1']
    elif isinstance(row[col + '_df1'], str) and isinstance(row[col + '_df2'], str) and row[col + '_df1'].lower() == row[col + '_df2'].lower():
        return row[col + '_df1']
    else:
        print("Alert:", col, row[col + '_df1'], row[col + '_df2'])
        return 'Alert'


def remove_collected_twice(df1, df2):
    # dedup
    df1_post_ids = df1["post_id"].tolist()
    df2_post_ids = df2["post_id"].tolist()

    # collected_twice = Same post_id with different collection_time 
    collected_twice = []
    for i in range(len(df1_post_ids)):
        if df1_post_ids[i] in df2_post_ids:
            tmp_1 = df1[df1['post_id'] == df1_post_ids[i]]
            tmp_2 = df2[df2['post_id'] == df1_post_ids[i]]
            collect_t_1 = tmp_1["collection_time"].values[0]
            collect_t_2 = tmp_2["collection_time"].values[0]
            if collect_t_1 != collect_t_2:
                collected_twice.append(df1_post_ids[i])

    print("posts that has been collected twice: (We delete its occurrence in the first df)\n", collected_twice)
    df1 = df1[~df1['post_id'].isin(collected_twice)]
    return df1

# not relevant as long as we delete one copy
def find_conflicts(merged_df, df1, df2, shared_cols):
    # Collects all conflicts with their metadata to report - based on combine_values 
    conflicts = {}
    for index, row in merged_df.iterrows():
        for col in shared_cols:
            if row[col] == "Alert" and col != "post_text":
                print(f"post: {row['post_id']}, column: {col}")
                conflict_post_id = str(row["post_id"])
                value_in_df1 = df1.loc[df1['post_id'] == conflict_post_id, col].values[0]
                value_in_df2 = df2.loc[df2['post_id'] == conflict_post_id, col].values[0]
                conflicts[conflict_post_id] = [col, value_in_df1, type(value_in_df1), value_in_df2, type(value_in_df2)]
    return conflicts


def reorder_cols(col_order, merged_df):
    final_col_order = []
    all_columns = merged_df.columns
    for col in col_order:
        if col in all_columns:
            final_col_order.append(col)
    return merged_df[final_col_order + [col for col in merged_df.columns if col not in final_col_order]]


# merge tables into one (ex: after_verification, all_data)
def merge_into_all(cls_task:ClassificationTask, df1, df2):
    shared_cols = [col for col in df2 if col in df1]
    shared_cols.remove("post_id")

    # convert numbers to string to avoid changes
    for col in ["post_id", "uploader_id"]:
        if col in df1.columns:
            df1[col] = df1[col].astype(str)
        if col in df2.columns:
            df2[col] = df2[col].astype(str)

    merged_df = pd.merge(df1, df2, on="post_id", how='outer', suffixes=('_df1', '_df2'))
    for col in shared_cols:
        merged_df[col] = merged_df.apply(lambda r: combine_values(r, col), axis=1)
        merged_df = merged_df.drop([f"{col}_df1", f"{col}_df2"], axis=1)

    # convert numbers to string to avoid changes
    for col in ["post_id", "uploader_id"]:
        if col in merged_df.columns:
            merged_df[col] = merged_df[col].astype(str)

    # reorder columns
    col_order = ["country", "username", "post_id", "uploader_id", "collection_time", "upload_date", "post_text",
                 "comments_amount", "retweet_amount", "like_amount", "views_amount", cls_task.get_property_col_name()]
    merged_df = reorder_cols(col_order, merged_df)

    # find conflicts' information
    conflicts = find_conflicts(merged_df, df1, df2, shared_cols)
    # property_col like: conflict_sentiment - for those rows not passing embs_similarity_threshold --> fillna None --> 0
    merged_df[cls_task.get_property_col_name()] = merged_df[cls_task.get_property_col_name()].fillna(0)

    # fill missing verified values
    # verification_col like: conflict_sentiment_v - some of the rows are filled by Kaman --> fill other rows eq to their property col i.e conflict_sentiment
    if cls_task.get_property_col_name() in merged_df.columns and cls_task.get_for_verification_col_name() in merged_df.columns:
        merged_df[cls_task.get_for_verification_col_name()] = merged_df[cls_task.get_for_verification_col_name()].fillna(merged_df[cls_task.get_property_col_name()])

    # Reports conflicts if any and throw - meaning no modifications to Storage, and can fix manually and restart merge
    if len(conflicts) != 0:  # there are conflicts
        error_message = f"Here is a list of conflict (post id: [columns, first value, second value]):\n {conflicts}" \
                        f"\nresolve conflicts and try again"
        raise ValueError(error_message)
    return merged_df


def to_file(df, file_name, suffix):
    file_created_time = str(datetime.now())
    current_date_time = "__updated_to_" + file_created_time
    current_date_time = current_date_time.split(".")[0].replace(" ", "__")
    current_date_time = current_date_time[:-3]
    current_date_time = current_date_time.replace(":", "_")
    file_name = file_name + current_date_time + suffix

    columns_to_save_as_text = ["post_id", "uploader_id"]
    uploaded_file_name = create_file(df, file_name, columns_to_save_as_text)
    return uploaded_file_name, file_created_time


def create_file(merged_data, file_name, columns_to_save_as_text=None):
    # Create a new Excel workbook and add a worksheet
    # Problem: post_id was corrupted (before change +'_a') 
    # Solution: Use Excel api to create workbook
    wb = Workbook()
    ws = wb.active

    # Define a cell format/style to set the cell as text
    text_format = NamedStyle(name='text_style', number_format='@')
    wb.add_named_style(text_format)

    # Write column names to the first row
    for col_idx, col_name in enumerate(merged_data.columns):
        cell = ws.cell(row=1, column=col_idx + 1, value=col_name)

    # Write the DataFrame to Excel with the specified text cell format/style for the selected columns
    for col_idx, col_name in enumerate(merged_data.columns):
        for row_idx, value in enumerate(merged_data[col_name]):
            cell = ws.cell(row=row_idx + 2, column=col_idx + 1, value=value)

            # Set the text cell format for the specified columns
            if col_name in columns_to_save_as_text:
                cell.style = 'text_style'

    # Save the workbook to a file
    wb.save(file_name)
    return file_name


def download_and_merge_files(cls_task: ClassificationTask, id_folder, all_data_one_country):
    # download all files from specific country to merge
    downloaded_country_files_lst = d.download_from_drive(parent_folder_id=id_folder, download_path=d.current_dir)
    print("finished downloading, start merging.")
    for file_name in downloaded_country_files_lst:
        cur_file = read_file(file_name)
        # dedup many after_verification files againt a single all-data file
        changed_file = remove_collected_twice(cur_file, all_data_one_country)
        all_data_one_country = merge_into_all(cls_task,changed_file, all_data_one_country)
    return all_data_one_country, downloaded_country_files_lst


def delete_local_file(path):
    try:
        if os.path.exists(path):
            os.remove(path)
    except PermissionError:
        print(f"File {path} is in use. Will retry after a short delay.")
        time.sleep(1)
        delete_local_file(path)


def run_whole_process(task_name=None,country=None,merge_from_folder=None):
    if country is None:
        countries = ["france", "germany", "uk", "us"]
        country_number = int(input("insert country's index: 0 - france, 1 - germany, 2 - uk, 3 - us\n"))
        country = countries[country_number]

    if task_name is None:
        tasks = [TaskName.IsraelSentimentTweet, TaskName.HouthisTweet, TaskName.Unrwa]
        task_number = int(input("insert task : 0 - " + TaskName.IsraelSentimentTweet.name + ", 1 - " + TaskName.HouthisTweet.name + ", 2 - " + TaskName.Unrwa.name + "\n"))
        task_name = tasks[task_number]
    cls_task = ClassificationTask(task_name=task_name)

    all_data_id_folder = IDs.get_folder_id(country, task_name, IDs.ALL_DATA_FOLDER_ID_DICT)

    all_data_file_name = d.download_from_drive(parent_folder_id=all_data_id_folder,
                                               download_path=d.current_dir)

    all_data_file_name = all_data_file_name[0]
    country_all_data_df = read_file(all_data_file_name)

    if merge_from_folder is None:
        merge_from_folder = int(
            input("\ninsert 0 to merge after_verification, 1 to merge other queries (deltas)\n"))
    if merge_from_folder == 0:
        after_ver_id_folder = IDs.get_folder_id(country, task_name, IDs.AFTER_VER_FOLDER_ID_DICT)

        final_df, downloaded_files_lst = download_and_merge_files(cls_task, after_ver_id_folder,
                                                                  country_all_data_df)
        origin_folder_id = after_ver_id_folder
    else:  # merge_from_folder == 1:
        delta_id_folder = IDs.get_folder_id(country, task_name, IDs.DELTA_FOLDER_ID_DICT)

        final_df, downloaded_files_lst = download_and_merge_files(cls_task, delta_id_folder,
                                                                  country_all_data_df)
        origin_folder_id = delta_id_folder

    print("finished merging, creating .xlsx file")
    file_name = f"all_data_{country}_{task_name.name}"
    suffix = ".xlsx"

    # create an Excel file
    uploaded_file_name, file_created_time = to_file(final_df, file_name, suffix)

    # upload file to google drive
    d.upload_to_drive(uploaded_file_name, all_data_id_folder)

    # Move all files from Data/after_verifications or Data/Deltas + Data/all_data/us/all_data_us__updated_to_<last_old_time> --> to Data/archive
    archive_id_folder = IDs.get_folder_id(country, task_name, IDs.ARCHIVE_FOLDER_ID_DICT)

    d.move_file_to_different_folder(all_data_file_name, all_data_id_folder,
                                    archive_id_folder)

    for file in downloaded_files_lst:
        d.move_file_to_different_folder(file, origin_folder_id, archive_id_folder)

    # delete local downloaded files
    total_downloaded_files = [all_data_file_name] + downloaded_files_lst + [uploaded_file_name]
    for file in total_downloaded_files:
        delete_local_file(file)

    # write to log3

    log_path = "./data/twitter_data/merge/log.txt"
    Path(log_path).parent.mkdir(parents=True,exist_ok=True)
    with open(log_path, 'a') as file:
        file.write(f"At: {file_created_time}\n"
                   f"the files: {downloaded_files_lst}\n"
                   f"was merged into the file: {all_data_file_name}\n\n")


def main():
    run_whole_process()


if __name__ == '__main__':
    main()
